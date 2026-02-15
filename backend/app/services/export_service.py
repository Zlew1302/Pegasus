"""Export service — generates CSV, Excel, and PDF reports for projects."""

import csv
import io
import logging
from datetime import datetime

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.project import Project
from app.models.task import Task

logger = logging.getLogger(__name__)


async def _get_project_tasks(db: AsyncSession, project_id: str):
    """Lade Projekt und alle Tasks."""
    proj_result = await db.execute(select(Project).where(Project.id == project_id))
    project = proj_result.scalar_one_or_none()

    task_result = await db.execute(
        select(Task)
        .where(Task.project_id == project_id)
        .order_by(Task.sort_order, Task.created_at)
    )
    tasks = task_result.scalars().all()

    return project, tasks


HEADERS = [
    "ID", "Titel", "Status", "Priorität", "Typ", "Beschreibung",
    "Akzeptanzkriterien", "Geschätzt (Min)", "Tatsächlich (Min)",
    "Deadline", "Erstellt", "Aktualisiert",
]

STATUS_DE = {
    "backlog": "Backlog",
    "todo": "Zu erledigen",
    "in_progress": "In Bearbeitung",
    "review": "Review",
    "done": "Erledigt",
    "blocked": "Blockiert",
}

PRIORITY_DE = {
    "critical": "Kritisch",
    "high": "Hoch",
    "medium": "Mittel",
    "low": "Niedrig",
}


def _task_row(task: Task) -> list[str]:
    return [
        task.id,
        task.title or "",
        STATUS_DE.get(task.status, task.status),
        PRIORITY_DE.get(task.priority, task.priority),
        task.task_type or "",
        (task.description or "")[:200],
        (task.acceptance_criteria or "")[:200],
        str(task.estimated_duration_minutes or ""),
        str(task.actual_duration_minutes or ""),
        task.deadline.isoformat() if task.deadline else "",
        task.created_at.isoformat() if task.created_at else "",
        task.updated_at.isoformat() if task.updated_at else "",
    ]


async def export_project_csv(db: AsyncSession, project_id: str) -> bytes:
    """Exportiert ein Projekt als CSV."""
    project, tasks = await _get_project_tasks(db, project_id)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(HEADERS)
    for task in tasks:
        writer.writerow(_task_row(task))

    return output.getvalue().encode("utf-8-sig")


async def export_project_excel(db: AsyncSession, project_id: str) -> bytes:
    """Exportiert ein Projekt als Excel-Datei."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    project, tasks = await _get_project_tasks(db, project_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E8590C", end_color="E8590C", fill_type="solid")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, task in enumerate(tasks, 2):
        for col_idx, value in enumerate(_task_row(task), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Project info sheet
    info_ws = wb.create_sheet("Projekt-Info")
    info_data = [
        ("Projekt", project.title if project else ""),
        ("Beschreibung", project.description if project else ""),
        ("Ziel", project.goal if project else ""),
        ("Status", project.status if project else ""),
        ("Anzahl Tasks", str(len(tasks))),
        ("Exportiert am", datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]
    for row_idx, (label, value) in enumerate(info_data, 1):
        info_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        info_ws.cell(row=row_idx, column=2, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_project_pdf(db: AsyncSession, project_id: str) -> bytes:
    """Exportiert ein Projekt als PDF-Report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    project, tasks = await _get_project_tasks(db, project_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "ProjTitle", parent=styles["Title"], fontSize=18, spaceAfter=6
    )
    heading_style = ParagraphStyle(
        "ProjHeading", parent=styles["Heading2"], fontSize=13, spaceAfter=4
    )

    elements = []

    # Title
    elements.append(Paragraph(f"Projekt-Report: {project.title if project else 'Unbekannt'}", title_style))
    elements.append(Spacer(1, 4 * mm))

    # Project info
    if project:
        info_text = []
        if project.description:
            info_text.append(f"<b>Beschreibung:</b> {project.description[:300]}")
        if project.goal:
            info_text.append(f"<b>Ziel:</b> {project.goal[:200]}")
        info_text.append(f"<b>Status:</b> {project.status}")
        info_text.append(f"<b>Anzahl Tasks:</b> {len(tasks)}")
        for t in info_text:
            elements.append(Paragraph(t, styles["Normal"]))
        elements.append(Spacer(1, 6 * mm))

    # Task table
    elements.append(Paragraph("Aufgaben", heading_style))

    # Simplified columns for PDF
    pdf_headers = ["Titel", "Status", "Priorität", "Typ", "Geschätzt", "Deadline"]
    table_data = [pdf_headers]
    for task in tasks:
        table_data.append([
            (task.title or "")[:40],
            STATUS_DE.get(task.status, task.status),
            PRIORITY_DE.get(task.priority, task.priority),
            task.task_type or "-",
            f"{task.estimated_duration_minutes} Min" if task.estimated_duration_minutes else "-",
            task.deadline.strftime("%d.%m.%Y") if task.deadline else "-",
        ])

    if len(table_data) > 1:
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8590C")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Keine Aufgaben vorhanden.", styles["Normal"]))

    # Footer
    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph(
        f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')} — Pegasus",
        styles["Normal"],
    ))

    doc.build(elements)
    return buf.getvalue()
